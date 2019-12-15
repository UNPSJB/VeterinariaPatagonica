from django.shortcuts import render
from django.template import loader, RequestContext
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required, permission_required
from django.urls import reverse
from .models import Cliente
from .forms import ClienteFormFactory, FiltradoForm
from VeterinariaPatagonica import tools
from .gestionDeClientes import *
from VeterinariaPatagonica.tools import GestorListadoQuerySet

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Importamos settings para poder tener a la mano la ruta de la carpeta media
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from django.views.generic import View
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import cm
from reportlab.lib import colors

clientesFiltrados = [] 

def clientes(request):
    context = {}#Defino el contexto.
    template = loader.get_template('GestionDeClientes/GestionDeClientes.html')#Cargo el template desde la carpeta templates/GestionDeClientes.
    return HttpResponse(template.render(context, request))#Devuelvo la url con el template armado.

def menuVer(usuario, cliente):

    menu = [[],[],[]]

    if usuario.has_perm("GestionDeClientes.cliente_modificar"):
        menu[0].append( (reverse("clientes:clienteModificar", args=(cliente.id,)), "Modificar cliente") )
        if cliente.baja:
            menu[0].append( (reverse("clientes:clienteHabilitar", args=(cliente.id,)), "Habilitar cliente") )
        else:
            menu[0].append( (reverse("clientes:clienteDeshabilitar", args=(cliente.id,)), "Deshabilitar cliente") )

    if usuario.has_perm("GestionDeClientes.cliente_listar_habilitados"):
        menu[1].append( (reverse("clientes:clienteVerHabilitados"), "Listar clientes habilitados") )
    if usuario.has_perm("GestionDeClientes.cliente_listar_no_habilitados"):
        menu[1].append( (reverse("clientes:clienteVerDeshabilitados"), "Listar clientes deshabilitados") )

    if usuario.has_perm("GestionDeClientes.cliente_crear"):
        menu[2].append( (reverse("clientes:clienteCrear"), "Crear cliente") )

    return [ item for item in menu if len(item) ]

def menuListar(usuario, habilitados):
    menu = [[],[]]

    if (not habilitados) and usuario.has_perm("GestionDeClientes.cliente_ver_habilitados"):
        menu[0].append( (reverse("clientes:clienteVerHabilitados"), "Listar clientes habilitados") )
    if habilitados and usuario.has_perm("GestionDeClientes.cliente_ver_no_habilitados"):
        menu[0].append( (reverse("clientes:clienteVerDeshabilitados"), "Listar clienets deshabilitados") )

    if usuario.has_perm("GestionDeClientes.cliente_crear"):
        menu[1].append( (reverse("clientes:clienteCrear"), "Crear Cliente") )
    return [ item for item in menu if len(item) ]

def menuModificar(usuario, cliente):

    menu = [[],[],[],[]]

    menu[0].append( (reverse("clientes:clienteVer", args=(cliente.id,)), "Ver cliente") )

    if cliente.baja:
        menu[1].append( (reverse("clientes:clienteHabilitar", args=(cliente.id,)), "Habilitar cliente") )
    else:
        menu[1].append( (reverse("clientes:clienteDeshabilitar", args=(cliente.id,)), "Deshabilitar cliente") )

    if usuario.has_perm("GestionDeClientes.cliente_listar_habilitados"):
        menu[2].append( (reverse("clientes:clienteVerHabilitados"), "Listar clientes habilitados") )
    if usuario.has_perm("GestionDeClientes.cliente_listar_no_habilitados"):
        menu[2].append( (reverse("clientes:clienteVerDeshabilitados"), "Listar clientes deshabilitados") )

    if usuario.has_perm("GestionDeClientes.cliente_crear"):
        menu[3].append( (reverse("clientes:clienteCrear"), "Crear cliente") )

    return [ item for item in menu if len(item) ]

def menuCrear(usuario, cliente):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeClientes.cliente_listar_habilitados"):
        menu[0].append( (reverse("clientes:clienteVerHabilitados"), "Listar clientes habilitados") )
    if usuario.has_perm("GestionDeClientes.cliente_listar_no_habilitados"):
        menu[0].append( (reverse("clientes:clienteVerDeshabilitados"), "Listar clientes deshabilitados") )

    return [ item for item in menu if len(item) ]

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeClientes.add_Cliente', raise_exception=True)
def modificar(request, id= None, irAMascotas=1): #irAMascotas=1 -> False, irAMasotas=0 -> True

    cliente = Cliente.objects.get(id=id) if id is not None else None
    ClienteForm = ClienteFormFactory(cliente)

    if (id==None):
        context = {"titulo": 1, 'usuario': request.user}
    else:
        context = {"titulo": 2, 'usuario': request.user}

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            cliente = form.save()
            return HttpResponseRedirect("/GestionDeClientes/ver/{}".format(cliente.id))
        else:
            context['form'] = form
            context['menu'] = menuModificar(request.user, cliente)
    else:
        context['form'] = ClienteForm(instance=cliente)
        context['menu'] = menuCrear(request.user, cliente)
    template = loader.get_template('GestionDeClientes/formulario.html')
    return HttpResponse(template.render(context, request))

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeClientes.delete_Cliente', raise_exception=True)
def habilitar(request, id):
    try:
        cliente = Cliente.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    cliente.baja = False
    cliente.save()

    return HttpResponseRedirect( "/GestionDeClientes/verHabilitados/" )

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeCliente.delete_Cliente', raise_exception=True)
def deshabilitar(request, id):

    try:
        cliente = Cliente.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    cliente.baja = True
    cliente.save()

    return HttpResponseRedirect( "/GestionDeClientes/verDeshabilitados/" )

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeClientes.delete_Cliente', raise_exception=True)
def eliminar(request, id):
    try:
        cliente = Cliente.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()
    if request.method == 'POST':
        cliente.delete()
        return HttpResponseRedirect ("/GestionDeClientes/verDeshabilitados/" )
    else:
        template = loader.get_template('GestionDeClientes/eliminar.html')
        context = {
            'usuario' : request.user,
            'id' : id
        }
        return HttpResponse (template.render (context, request))

def ver(request, id):

    try:
        cliente = Cliente.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404("No encontrado", "El cliente con id={} no existe.".format(id))

    template = loader.get_template('GestionDeClientes/ver.html')
    contexto = {
        'cliente': cliente,
        'usuario': request.user,
        "menu" : menuVer(request.user, cliente)
    }

    return HttpResponse(template.render(contexto, request))

def verHabilitados(request, habilitados=True):
    """ Listado de clientes habilitados """

    clientes = Cliente.objects.habilitados()

    gestor = GestorListadoQuerySet(
        campos=[
            ["orden_dniCuit", "DNI/CUIT"],
            ["orden_apellidos", "Apellidos"],
            ["orden_nombres", "Nombres"],
            ["orden_localidad", "Localidad"],
            ["orden_tipoDeCliente", "Tipo De Cliente"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=clientes,
        mapaFiltrado= Cliente.MAPPER,
        mapaOrden= clientes.MAPEO_ORDEN
    )
   
    gestor.cargar(request)

    template = loader.get_template('GestionDeClientes/verHabilitados.html')
    context = {
        "gestor" : gestor, 
        "menu" : menuListar(request.user, habilitados),
    }
    return HttpResponse(template.render ( context, request ))


def verDeshabilitados(request, habilitados=False):
    """ Listado de clientes deshabilitados """
    global clientesFiltrados

    clientes = Cliente.objects.deshabilitados()

    gestor = GestorListadoQuerySet(
        campos=[
            ["orden_dniCuit", "DNI/CUIT"],
            ["orden_apellidos", "Apellidos"],
            ["orden_nombres", "Nombres"],
            ["orden_localidad", "Localidad"],
            ["orden_tipoDeCliente", "Tipo De Cliente"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=clientes,
        mapaFiltrado= Cliente.MAPPER,
        mapaOrden= clientes.MAPEO_ORDEN
    )

    
    gestor.cargar(request)

    clientesFiltrados = gestor.queryset
    template = loader.get_template('GestionDeClientes/verDeshabilitados.html')
    context = {
        "gestor" : gestor,
        "menu" : menuListar(request.user, habilitados),}
    return HttpResponse (template.render (context, request))

def ListadoClientesExcel(request):
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    ws = wb.worksheets[0]
    # En la celda B1 ponemos el texto 'LISTADO DE CLIENTES'
    ws['B1'] = 'LISTADO DE CLIENTES'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:F1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'DNI/CUIT'
    ws['C3'] = 'NOMBRES'
    ws['D3'] = 'APELLIDOS'
    ws['E3'] = 'LOCALIDAD'
    ws['F3'] = 'TIPO DE CLIENTE'
    cont = 5
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for cliente in clientesFiltrados:
        ws.cell(row=cont, column=2).value = cliente.dniCuit
        ws.cell(row=cont, column=3).value = cliente.nombres
        ws.cell(row=cont, column=4).value = cliente.apellidos
        ws.cell(row=cont, column=5).value = cliente.localidad
        ws.cell(row=cont, column=6).value = cliente.tipoDeCliente
        cont = cont + 1

    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]

    for i, column_width in enumerate(column_widths):
         ws.column_dimensions[get_column_letter(i+1)].width = column_width

    # Establecemos el nombre del archivo
    nombre_archivo = "ListadoClientes.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

def ListadoClientesPDF(request):
    print("GET")
    # Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    # La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    # Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = canvas.Canvas(buffer)
    # Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera(pdf)
    y = 500
    tabla(pdf, y, clientesFiltrados)
    # Con show page hacemos un corte de página para pasar a la siguiente
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def cabecera(pdf):
    print("CABECERA")
    # Utilizamos el archivo logo_vetpat.png que está guardado en la carpeta media/imagenes
    archivo_imagen = settings.MEDIA_ROOT + '/imagenes/logo_vetpat2.jpeg'
    # Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 20, 750, 120, 90, preserveAspectRatio=True)
    # Establecemos el tamaño de letra en 16 y el tipo de letra Helvetica
    pdf.setFont("Helvetica", 16)
    # Dibujamos una cadena en la ubicación X,Y especificada
    pdf.drawString(190, 790, u"VETERINARIA PATAGONICA")
    pdf.setFont("Helvetica", 14)
    pdf.drawString(220, 770, u"LISTADO DE CLIENTES")

def tabla(pdf, y, clientes):
    print("TABLA")
    # Creamos una tupla de encabezados para neustra tabla
    encabezados = ('DNI/CUIT', 'Nombres', 'Apellidos', 'Localidad', 'Tipo de Cliente')
    # Creamos una lista de tuplas que van a contener a las personas
    detalles = [(cliente.dniCuit, cliente.nombres, cliente.apellidos, cliente.localidad, cliente.tipoDeCliente) for cliente in clientes]
    # Establecemos el tamaño de cada una de las columnas de la tabla
    detalle_orden = Table([encabezados] + detalles, colWidths=[3 * cm, 5 * cm, 5 * cm, 4 * cm, 3 * cm])
    # Aplicamos estilos a las celdas de la tabla
    detalle_orden.setStyle(TableStyle(
        [
            # La primera fila(encabezados) va a estar centrada
            ('ALIGN', (0, 0), (3, 0), 'CENTER'),
            # Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            # El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]
    ))
    # Establecemos el tamaño de la hoja que ocupará la tabla
    detalle_orden.wrapOn(pdf, 800, 600)
    # Definimos la coordenada donde se dibujará la tabla
    detalle_orden.drawOn(pdf, 20, y)

@login_required
def ayudaContextualCliente(request):
# Redireccionamos la ayuda contextual
    template = loader.get_template('GestionDeClientes/ayudaContextualCliente.html')
    contexto = {
        'usuario': request.user,
    }
    return HttpResponse(template.render(contexto, request))
