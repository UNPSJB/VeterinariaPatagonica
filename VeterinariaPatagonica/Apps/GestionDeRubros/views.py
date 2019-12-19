# Create your views here.

from django.shortcuts import render
from django.template import loader
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required, permission_required
from .models import Rubro
from .forms import RubroFormFactory, FiltradoForm
from VeterinariaPatagonica import tools
from django.db.models import Q
from django.urls import reverse_lazy
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from VeterinariaPatagonica.tools import GestorListadoQuerySet


#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
#Importamos settings para poder tener a la mano la ruta de la carpeta media
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from django.views.generic import View
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import cm
from reportlab.lib import colors

rubrosFiltrados = []

def rubros(request):
    context = {}#Defino el contexto.
    template = loader.get_template('GestionDeRubros/GestionDeRubros.html')#Cargo el template desde la carpeta templates/GestionDeRubros.
    return HttpResponse(template.render(context, request))#Devuelvo la url con el template armado.

@permission_required('GestionDeRubros.deshabilitar_rubro', raise_exception=True)
def dispatch(self, *args, **kwargs):
    return super(Rubro, self).dispatch(*args, **kwargs)

def menuVer(usuario, rubro):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeRubros.rubro_modificar"):
        menu[0].append( (reverse("rubros:rubroModificar", args=(rubro.id,)), "Modificar rubro") )
        if rubro.baja:
            menu[0].append( (reverse("rubros:rubroHabilitar", args=(rubro.id,)), "Habilitar rubro") )
        else:
            menu[0].append( (reverse("rubros:rubroDeshabilitar", args=(rubro.id,)), "Deshabilitar rubro") )

    if usuario.has_perm("GestionDeRubros.rubro_listar_habilitados"):
        menu[1].append( (reverse("rubros:rubroVerHabilitados"), "Listar rubros habilitados") )
    if usuario.has_perm("GestionDeRubros.rubro_listar_no_habilitados"):
        menu[1].append( (reverse("rubros:rubroVerDeshabilitados"), "Listar rubros deshabilitados") )

    if usuario.has_perm("GestionDeRubros.rubro_crear"):
        menu[2].append( (reverse("rubros:rubroCrear"), "Crear rubro") )


    return [ item for item in menu if len(item) ]

def menuListar(usuario, habilitados):
    menu = [[], [], [], [], []]

    if (not habilitados) and usuario.has_perm("GestionDeRubros.rubro_ver_habilitados"):

        menu[0].append( (reverse("rubros:rubroVerHabilitados"), "Listar rubros habilitados") )
        menu[1].append( (reverse("rubros:rubrosListadoExcel"), "Exportar rubros deshabilitados"))
        menu[2].append( (reverse("rubros:rubrosListadoPDF"), "Imprimir rubros deshabilitados"))

    if habilitados and usuario.has_perm("GestionDeRubros.rubro_ver_no_habilitados"):
        menu[0].append( (reverse("rubros:rubroVerDeshabilitados"), "Listar rubros deshabilitados") )
        menu[1].append( (reverse("rubros:rubrosListadoExcel"), "Exportar rubros habilitados") )
        menu[2].append( (reverse("rubros:rubrosListadoPDF"), "Imprimir rubros habilitados") )

    if usuario.has_perm("GestionDeRubros.rubro_crear"):
        menu[3].append( (reverse("rubros:rubroCrear"), "Crear Rubro") )


    return [ item for item in menu if len(item) ]

def menuModificar(usuario, rubro):

    menu = [[],[],[],[],[]]

    menu[0].append( (reverse("rubros:rubroVer", args=(rubro.id,)), "Ver rubro") )

    if rubro.baja:
        menu[1].append( (reverse("rubros:rubroHabilitar", args=(rubro.id,)), "Habilitar rubro") )
    else:
        menu[1].append( (reverse("rubros:rubroDeshabilitar", args=(rubro.id,)), "Deshabilitar rubro") )

    if usuario.has_perm("GestionDeRubros.rubro_listar_habilitados"):
        menu[2].append( (reverse("rubros:rubroVerHabilitados"), "Listar rubros habilitados") )
    if usuario.has_perm("GestionDeRubros.rubro_listar_no_habilitados"):
        menu[2].append( (reverse("rubros:rubroVerDeshabilitados"), "Listar rubros deshabilitados") )

    if usuario.has_perm("GestionDeRubros.rubro_crear"):
        menu[3].append( (reverse("rubros:rubroCrear"), "Crear rubro") )



    return [ item for item in menu if len(item) ]

def menuCrear(usuario, rubro):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeRubros.rubro_listar_habilitados"):
        menu[0].append( (reverse("rubros:rubroVerHabilitados"), "Listar rubros habilitados") )
    if usuario.has_perm("GestionDeRubros.rubro_listar_no_habilitados"):
        menu[0].append( (reverse("rubros:rubroVerDeshabilitados"), "Listar rubros deshabilitados") )

    return [ item for item in menu if len(item) ]


@login_required(redirect_field_name='proxima')
@permission_required('GestionDeRubros.add_Rubro', raise_exception=True)
def modificar(request, id = None):

    rubro = Rubro.objects.get(id=id) if id is not None else None
    RubroForm = RubroFormFactory(rubro)
    if (id==None):
        context = {"titulo": 1, 'usuario': request.user}
    else:
        context = {"titulo": 2, 'usuario': request.user}
    if request.method == 'POST':
        formulario = RubroForm(request.POST, instance=rubro)
        if formulario.is_valid():
            rubro = formulario.save()
            return HttpResponseRedirect("/GestionDeRubros/ver/{}".format(rubro.id))
        else:
            context['form'] = formulario
            context['menu'] = menuModificar(request.user, rubro)
    else:
        context['form'] = RubroForm(instance=rubro)
        context['menu'] = menuCrear(request.user, rubro)
    template = loader.get_template('GestionDeRubros/formulario.html')
    return HttpResponse(template.render(context, request))


@login_required(redirect_field_name='proxima')
@permission_required('GestionDeRubros.delete_Rubro', raise_exception=True)
def habilitar(request, id):
    try:
        rubro = Rubro.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    rubro.baja = False
    rubro.save()

    return HttpResponseRedirect( "/GestionDeRubros/verHabilitados/" )

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeRubros.delete_Rubro', raise_exception=True)
def deshabilitar(request, id):

    try:
        rubro = Rubro.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    rubro.baja = True
    rubro.save()

    return HttpResponseRedirect( "/GestionDeRubros/verDeshabilitados/" )

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeRubros.delete_Rubro', raise_exception=True)
def eliminar(request, id):
    try:
        rubro = Rubro.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()
    if request.method == 'POST':
        rubro.delete()
        return HttpResponseRedirect( "/GestionDeRubros/verDeshabilitados/" )
    else:
        template = loader.get_template('GestionDeRubros/eliminar.html')
        context = {
            'usuario' : request.user,
            'id' : id
        }
        return HttpResponse( template.render( context, request) )

def ver(request, id):

    try:
        rubro = Rubro.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404("No encontrado", "El rubro con id={} no existe.".format(id))

    template = loader.get_template('GestionDeRubros/ver.html')
    contexto = {
        'rubro': rubro,
        'usuario': request.user,
        "menu" : menuVer(request.user, rubro)
    }

    return HttpResponse(template.render(contexto, request))

def verHabilitados(request, habilitados=True):
    global rubrosFiltrados
    rubros = Rubro.objects.habilitados()
    gestor = GestorListadoQuerySet(
        campos=[
            ["orden_nombre", "Nombre"],
            ["orden_descripcion", "Descripcion"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=rubros,
        mapaFiltrado= Rubro.MAPPER,
        mapaOrden= rubros.MAPEO_ORDEN
    )
    gestor.cargar(request)
    rubrosFiltrados = gestor.queryset
    template = loader.get_template('GestionDeRubros/verHabilitados.html')
    context = {
        "gestor": gestor,
        "menu" : menuListar(request.user, habilitados)
    }

    return  HttpResponse(template.render(context,request))


def verDeshabilitados(request, habilitados=False):
    global rubrosFiltrados
    rubros = Rubro.objects.deshabilitados()
    gestor = GestorListadoQuerySet(
        campos=[
            ["orden_nombre", "Nombre"],
            ["orden_descripcion", "Descripcion"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=rubros,
        mapaFiltrado= Rubro.MAPPER,
        mapaOrden= rubros.MAPEO_ORDEN
    )
    gestor.cargar(request)
    rubrosFiltrados = gestor.queryset
    template = loader.get_template('GestionDeRubros/verDeshabilitados.html')
    context = {
        "gestor": gestor,
        "menu" : menuListar(request.user, habilitados),
    }
    return  HttpResponse(template.render(context,request))

@login_required
def ayudaContextualRubro(request):
# Redireccionamos la ayuda contextual
    template = loader.get_template('GestionDeRubros/ayudaContextualRubro.html')
    contexto = {
        'usuario': request.user,
    }
    return HttpResponse(template.render(contexto, request))


#LISTADOS
def ListadoRubrosExcel(request):
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    ws = wb.worksheets[0]
    # En la celda B1 ponemos el texto 'LISTADO DE CLIENTES'
    ws['B1'] = 'LISTADO DE RUBROS'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:F1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'DESCRIPCION'
    cont = 2
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for rubro in rubrosFiltrados:
        ws.cell(row=cont, column=2).value = rubro.nombre
        ws.cell(row=cont, column=3).value = rubro.descripcion
        cont = cont + 1

    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]

    for i, column_width in enumerate(column_widths):
         ws.column_dimensions[get_column_letter(i+1)].width = column_width

    # Establecemos el nombre del archivo
    nombre_archivo = "ListadoRubros.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

def ListadoRubrosPDF(request):
    print("GET")
    # Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    # La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    # Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = canvas.Canvas(buffer)
    # Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera(pdf)
    y = 700
    tabla(pdf, y, rubrosFiltrados)
    # Con show page hacemos un corte de página para pasar a la siguiente
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def cabecera(pdf):
    print("CABECERA")
    # Utilizamos el archivo logo_vetpat.png que está guardado en la carpeta media/imagenes
    archivo_imagen = settings.MEDIA_ROOT + '/imagenes/logo_vetpat2.jpeg'
    # Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 20, 750, 120, 90, preserveAspectRatio=True)
    # Establecemos el tamaño de letra en 16 y el tipo de letra Helvetica
    pdf.setFont("Helvetica", 16)
    # Dibujamos una cadena en la ubicación X,Y especificada
    pdf.drawString(190, 790, u"VETERINARIA PATAGÓNICA")
    pdf.setFont("Helvetica", 14)
    pdf.drawString(220, 770, u"LISTADO DE RUBROS")

def tabla(pdf, y, rubros):
    print("TABLA")
    # Creamos una tupla de encabezados para neustra tabla
    encabezados = ('Nombre', 'Descripcion')
    # Creamos una lista de tuplas que van a contener a las personas
    detalles = []
    for rubro in rubros:
        y -= 20
        r = (rubro.nombre, rubro.descripcion)
        detalles.append(r)
    # Establecemos el tamaño de cada una de las columnas de la tabla
    detalle_orden = Table([encabezados] + detalles, colWidths=[3 * cm,5 * cm])
    # Aplicamos estilos a las celdas de la tabla
    detalle_orden.setStyle(TableStyle(
        [
            # La primera fila(encabezados) va a estar centrada
            ('ALIGN', (0, 0), (3, 0), 'CENTER'),
            # Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            # El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]
    ))
    # Establecemos el tamaño de la hoja que ocupará la tabla
    detalle_orden.wrapOn(pdf, 800, 600)
    # Definimos la coordenada donde se dibujará la tabla
    detalle_orden.drawOn(pdf, 20, y)
    