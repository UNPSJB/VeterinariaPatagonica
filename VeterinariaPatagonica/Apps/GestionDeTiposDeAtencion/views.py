from os.path import join as pathjoin

from django.template import loader
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib.auth.models import AnonymousUser
from django.core.exceptions import PermissionDenied
from django import forms

from .models import TipoDeAtencion
from .forms import TipoDeAtencionForm, ModificacionTipoDeAtencionForm, FiltradoForm
from VeterinariaPatagonica.errores import VeterinariaPatagonicaError
from VeterinariaPatagonica.tools import GestorListadoQuerySet
from django.contrib.auth.decorators import login_required

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Importamos settings para poder tener a la mano la ruta de la carpeta media
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from django.views.generic import View
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import cm
from reportlab.lib import colors

tdasFiltrados = [] 


LOGIN_URL = '/login/'

def plantilla(*ruta):
    return pathjoin("GestionDeTiposDeAtencion", *ruta) + ".html"


def permisosModificar(tipoDeAtencion):
    permisos = ["GestionDeTiposDeAtencion.tipodeatencion_modificar"]
    if tipoDeAtencion.baja:
        permisos.append("GestionDeTiposDeAtencion.tipodeatencion_ver_no_habilitados")
    else:
        permisos.append("GestionDeTiposDeAtencion.tipodeatencion_ver_habilitados")
    return permisos


def menuModificar(usuario, tipoDeAtencion):

    menu = [[],[],[],[]]

    '''menu[0].append( (reverse("tiposDeAtencion:ver", args=(tipoDeAtencion.id,)), "Ver tipo de atención") )'''

    '''if tipoDeAtencion.baja:
        menu[1].append( (reverse("tiposDeAtencion:habilitar", args=(tipoDeAtencion.id,)), "Habilitar tipo de atención") )
    else:
        menu[1].append( (reverse("tiposDeAtencion:deshabilitar", args=(tipoDeAtencion.id,)), "Deshabilitar tipo de atención") )'''

    if usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_listar_habilitados"):
        menu[0].append( (reverse("tiposDeAtencion:habilitados"), "Listar tipos de atención habilitados") )
    if usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_listar_no_habilitados"):
        menu[1].append( (reverse("tiposDeAtencion:deshabilitados"), "Listar tipos de atención deshabilitados") )

    '''if usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_crear"):
        menu[3].append( (reverse("tiposDeAtencion:crear"), "Crear tipo de atención") )'''

    return [ item for item in menu if len(item) ]



def permisosVer(tipoDeAtencion):
    permisos = []
    if tipoDeAtencion.baja:
        permisos.append("GestionDeTiposDeAtencion.tipodeatencion_ver_no_habilitados")
    else:
        permisos.append("GestionDeTiposDeAtencion.tipodeatencion_ver_habilitados")
    return permisos



def menuVer(usuario, tipoDeAtencion):

    menu = [[],[],[]]

    if usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_modificar"):
        menu[0].append( (reverse("tiposDeAtencion:modificar", args=(tipoDeAtencion.id,)), "Modificar tipo de atención") )
        if tipoDeAtencion.baja:
            menu[0].append( (reverse("tiposDeAtencion:habilitar", args=(tipoDeAtencion.id,)), "Habilitar tipo de atención") )
        else:
            menu[0].append( (reverse("tiposDeAtencion:deshabilitar", args=(tipoDeAtencion.id,)), "Deshabilitar tipo de atención") )

    if usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_listar_habilitados"):
        menu[1].append( (reverse("tiposDeAtencion:habilitados"), "Listar tipos de atención habilitados") )
    if usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_listar_no_habilitados"):
        menu[1].append( (reverse("tiposDeAtencion:deshabilitados"), "Listar tipos de atención deshabilitados") )

    if usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_crear"):
        menu[2].append( (reverse("tiposDeAtencion:crear"), "Crear tipo de atención") )

    return [ item for item in menu if len(item) ]



def permisosListar(habilitados):
    permisos = []
    if habilitados:
        permisos.append("GestionDeTiposDeAtencion.tipodeatencion_listar_habilitados")
    else:
        permisos.append("GestionDeTiposDeAtencion.tipodeatencion_listar_no_habilitados")
    return permisos



def menuListar(usuario, habilitados):
    menu = [[],[],[],[]]

    if (not habilitados) and usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_listar_habilitados"):
        menu[0].append( (reverse("tiposDeAtencion:habilitados"), "Listar tipos de atención habilitados") )
        menu[1].append( (reverse("tiposDeAtencion:tdaListadoEXCEL"), "Exportar tipos de atención deshabilitados") )
        menu[2].append( (reverse("tiposDeAtencion:tdaListadoPDF"), "Imprimir tipos de atención deshabilitados") )

    if habilitados and usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_listar_no_habilitados"):
        menu[0].append( (reverse("tiposDeAtencion:deshabilitados"), "Listar tipos de atención deshabilitados") )
        menu[1].append( (reverse("tiposDeAtencion:tdaListadoEXCEL"), "Exportar tipos de atención habilitados") )
        menu[2].append( (reverse("tiposDeAtencion:tdaListadoPDF"), "Imprimir tipos de atención habilitados") )

    if usuario.has_perm("GestionDeTiposDeAtencion.tipodeatencion_crear"):
        menu[3].append( (reverse("tiposDeAtencion:crear"), "Crear tipo de atención") )
    return [ item for item in menu if len(item) ]



def listar(request, habilitados=True):

    global tdasFiltrados

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    if not request.user.has_perms(permisosListar(habilitados)):
        raise PermissionDenied()

    if habilitados:
        queryset = TipoDeAtencion.objects.habilitados()
    else:
        queryset = TipoDeAtencion.objects.deshabilitados()

    gestor = GestorListadoQuerySet(
        campos = [
            ["nombre", "Nombre"],
            ["emergencia", "Emergencia"],
            ["lugar", "Lugar de atencion"],
            ["recargo", "Recargo"],
            ["iniciofranjahoraria", "Franja Horaria"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=queryset,
        mapaFiltrado=queryset.MAPEO_FILTRADO,
        mapaOrden=queryset.MAPEO_ORDEN
    )

    gestor.cargar(request)
    gestor.paginacion["cantidad"].label = "Resultados por página"
    tdasFiltrados = gestor.queryset
    template = loader.get_template( plantilla("listar") )
    context = {
        "gestor" : gestor,
        "tipo" : "habilitados" if habilitados else "deshabilitados",
        "menu" : menuListar(request.user, habilitados),
    }
    return HttpResponse(template.render( context, request ))



def ver(request, id):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    try:
        tipoDeAtencion = TipoDeAtencion.objects.get(id=id)
    except TipoDeAtencion.DoesNotExist:
        raise VeterinariaPatagonicaError(
            "Error",
            "Tipo de atención no encontrado"
        )

    if not request.user.has_perms(permisosVer(tipoDeAtencion)):
        raise PermissionDenied()

    context = {
        "tipoDeAtencion" : tipoDeAtencion,
        "menu" : menuVer(request.user, tipoDeAtencion)
    }

    template = loader.get_template(plantilla("ver"))
    return HttpResponse(template.render( context, request ))



def crear(request):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    if not request.user.has_perm("GestionDeTiposDeAtencion.tipodeatencion_crear"):
        raise PermissionDenied()

    if request.method == "POST":
        tipoDeAtencion = TipoDeAtencion()
        form = TipoDeAtencionForm(request.POST, instance=tipoDeAtencion)

        if form.is_valid():
            tipoDeAtencion = form.save()
            return HttpResponseRedirect( reverse("tiposDeAtencion:ver", args=(tipoDeAtencion.id,)) )

    else:
        form = TipoDeAtencionForm()

    menu = [[]]
    if request.user.has_perm("GestionDeTiposDeAtencion.tipodeatencion_listar_habilitados"):
        menu[0].append( (reverse("tiposDeAtencion:habilitados"), "Listar tipos de atención habilitados") )
    if request.user.has_perm("GestionDeTiposDeAtencion.tipodeatencion_listar_no_habilitados"):
        menu[0].append( (reverse("tiposDeAtencion:deshabilitados"), "Listar tipos de atención deshabilitados") )

    context = {
        "accion" : "crear",
        "form" : form,
        "menu" : [ item for item in menu if len(item) ]
    }
    template = loader.get_template(plantilla("crear"))
    return HttpResponse(template.render( context, request) )



def modificar(request, id):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    try:
        tipoDeAtencion = TipoDeAtencion.objects.get(id=id)
    except TipoDeAtencion.DoesNotExist:
        raise VeterinariaPatagonicaError(
            "Error",
            "Tipo de atención no encontrado"
        )

    if not request.user.has_perms(permisosModificar(tipoDeAtencion)):
        raise PermissionDenied()

    if request.method == "POST":
        form = ModificacionTipoDeAtencionForm(request.POST, instance=tipoDeAtencion)

        if form.is_valid():
            if form.has_changed():
                form.save()
            return HttpResponseRedirect( reverse("tiposDeAtencion:ver", args=(tipoDeAtencion.id,)) )

    else:
        form = ModificacionTipoDeAtencionForm(instance=tipoDeAtencion)

    context = {
        "accion" : "modificar",
        "form" : form,
        "menu" : menuModificar(request.user, tipoDeAtencion)
    }

    template = loader.get_template(plantilla("modificar"))
    return HttpResponse(template.render( context, request) )



def cambioEstado(request, id, baja=False):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    try:
        tipoDeAtencion = TipoDeAtencion.objects.get(id=id)
    except TipoDeAtencion.DoesNotExist:
        raise VeterinariaPatagonicaError(
            "Error",
            "Tipo de atención no encontrado"
        )
        
    permisos = ["GestionDeTiposDeAtencion.tipodeatencion_modificar"]
    if tipoDeAtencion.baja:
        permisos.append("GestionDeTiposDeAtencion.tipodeatencion_ver_no_habilitados")
    else:
        permisos.append("GestionDeTiposDeAtencion.tipodeatencion_ver_habilitados")

    if not request.user.has_perms(permisos):
        raise PermissionDenied()

    if tipoDeAtencion.baja != baja:
        tipoDeAtencion.baja = baja
        tipoDeAtencion.save()

    return HttpResponseRedirect( reverse("tiposDeAtencion:ver", args=(tipoDeAtencion.id,)) )

def ListadoTdasExcel(request):
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    ws = wb.worksheets[0]
    # En la celda B1 ponemos el texto 'LISTADO DE CLIENTES'
    ws['B1'] = 'LISTADO DE TIPOS DE ATENCIÓN'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:F1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'EMERGENCIA'
    ws['D3'] = 'LUGAR'
    ws['E3'] = 'RECARGO'
    ws['F3'] = 'INICIO FRANJA HORARIA'
    ws['G3'] = 'FIN FRANJA HORARIA'
    cont = 5
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
 
    for tda in tdasFiltrados:
        ws.cell(row=cont, column=2).value = tda.nombre
        ws.cell(row=cont, column=3).value = tda.emergencia
        ws.cell(row=cont, column=4).value = tda.lugar
        ws.cell(row=cont, column=5).value = tda.recargo
        ws.cell(row=cont, column=6).value = tda.inicioFranjaHoraria
        ws.cell(row=cont, column=7).value = tda.finFranjaHoraria
        cont = cont + 1


    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]

    for i, column_width in enumerate(column_widths):
         ws.column_dimensions[get_column_letter(i+1)].width = column_width

    # Establecemos el nombre del archivo
    nombre_archivo = "ListadoTiposDeAtencion.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

def ListadoTdasPDF(request):
    print("GET")
    # Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    # La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    # Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = canvas.Canvas(buffer)
    # Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera(pdf)
    y = 700
    tabla(pdf, y, tdasFiltrados)
    # Con show page hacemos un corte de página para pasar a la siguiente
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def cabecera(pdf):
    print("CABECERA")
    # Utilizamos el archivo logo_vetpat.png que está guardado en la carpeta media/imagenes
    archivo_imagen = settings.MEDIA_ROOT + '/imagenes/logo_vetpat2.jpeg'
    # Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 20, 750, 120, 90, preserveAspectRatio=True)
    # Establecemos el tamaño de letra en 16 y el tipo de letra Helvetica
    pdf.setFont("Helvetica", 16)
    # Dibujamos una cadena en la ubicación X,Y especificada
    pdf.drawString(190, 790, u"VETERINARIA PATAGÓNICA")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(220, 770, u"LISTADO DE TIPOS DE ATENCIÓN")

def tabla(pdf, y, tiposDeAtencion):
    print("TABLA")
    # Creamos una tupla de encabezados para neustra tabla
    encabezados = ('Nombre', 'Emergencia', 'Lugar de atención', 'Recargo %', 'Inicio Franja Horaria', 'Fin Franja Horaria')
    # Creamos una lista de tuplas que van a contener a las personas

    detalles = []
    for tda in tiposDeAtencion:
        y -= 20
        t = (tda.nombre, tda.emergencia, tda.lugar, tda.recargo, tda.inicioFranjaHoraria, tda.finFranjaHoraria)
        detalles.append(t)
    
    # Establecemos el tamaño de cada una de las columnas de la tabla
    detalle_orden = Table([encabezados] + detalles, colWidths=[6.5 * cm, 2 * cm, 3 * cm, 2 * cm, 3 * cm, 3 * cm])
    # Aplicamos estilos a las celdas de la tabla
    detalle_orden.setStyle(TableStyle(
        [
            # La primera fila(encabezados) va a estar centrada
            ('ALIGN', (0, 0), (3, 0), 'CENTER'),
            # Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            # El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]
    ))
    # Establecemos el tamaño de la hoja que ocupará la tabla
    detalle_orden.wrapOn(pdf, 800, 600)
    # Definimos la coordenada donde se dibujará la tabla
    detalle_orden.drawOn(pdf, 30, y)


@login_required
def ayudaContextualTipoDeAtencion(request):

    template = loader.get_template('GestionDeTiposDeAtencion/ayudaContextualTiposdeAtencion.html')
    contexto = {
        'usuario': request.user,
    }

    return HttpResponse(template.render(contexto, request))