from datetime import datetime

from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.contrib.auth.models import AnonymousUser
from django.core.exceptions import PermissionDenied

from VeterinariaPatagonica.areas import Areas
from VeterinariaPatagonica.tools import guardarWorkbook
from .models.practica import Practica
from .models.estado import *
from .gestionDePracticas import *
from .forms import  FiltradoRealizacionesForm
from . import reportes

from openpyxl import Workbook


def listarXlsx(nombre, resultados, visibles, encabezados=None):

    libro = Workbook()
    hoja = libro.create_sheet(nombre, 0)
    fila = 1
    if encabezados:
        columna = 1
        for encabezado in encabezados:
            celda = hoja.cell(fila, columna, encabezado)
            columna += 1
        fila += 1
    for resultado in resultados:
        columna = 1
        if visibles[0]:
            celda = hoja.cell(fila, columna, resultado.id)
            columna += 1
        if visibles[1]:
            celda = hoja.cell(fila, columna, resultado.nombre_estado_actual)
            columna += 1
        if visibles[2]:
            celda = hoja.cell(fila, columna, str(resultado.cliente))
            columna += 1
        if visibles[3]:
            mascota = str(resultado.mascota) if resultado.mascota else ""
            celda = hoja.cell(fila, columna, mascota)
            columna += 1
        if visibles[4]:
            celda = hoja.cell(fila, columna, resultado.tipoDeAtencion.nombre)
            columna += 1
        if visibles[5]:
            celda = hoja.cell(fila, columna, resultado.precio)
            columna += 1
        if visibles[6]:
            celda = hoja.cell(fila, columna, resultado.marca_creacion)
            columna += 1
        if visibles[7]:
            celda = hoja.cell(fila, columna, resultado.marca_ultima_actualizacion)
            columna += 1
        if visibles[8]:
            celda = hoja.cell(fila, columna, resultado.nombre_creada_por)
            columna += 1
        if visibles[9]:
            celda = hoja.cell(fila, columna, resultado.nombre_atendida_por)
            columna += 1
        fila += 1
    return guardarWorkbook(libro)


def reporteHtml(area, practicas, hoy, dias):

    fecha = hoy - timedelta(days=dias["perfiles"])
    antes, despues = reportes.clasificarRealizaciones(practicas, fecha)
    datosAntes = reportes.datosPerfiles(
        reportes.contarRealizaciones(antes)
    )
    datosDespues = reportes.datosPerfiles(
        reportes.contarRealizaciones(despues)
    )
    perfiles = {
        "dias" : dias["perfiles"],
        "fecha" : fecha,
        "datos" : {
            "antes" : datosAntes,
            "despues" : datosDespues,
        },
    }

    fecha = hoy - timedelta(days=dias["realizaciones"])
    datos = reportes.datosRealizacionesPorDia(
        reportes.realizacionesEntre(practicas, fecha, hoy)
    )
    realizacionesPorDia = {
        "dias" : dias["realizaciones"],
        "fecha" : fecha,
        "datos" : datos,
    }

    fecha = hoy - timedelta(days=dias["actualizaciones"])
    seleccionadas = reportes.creadasEntre(practicas, fecha, hoy)
    porcentajesActualizacion = {
        "datos" : None,
        "dias" : dias["actualizaciones"],
        "fecha" : fecha,
    }
    if seleccionadas:
        datos = reportes.datosPorcentajesActualizacion(
            reportes.calcularNiveles(seleccionadas, area)
        )
        porcentajesActualizacion["datos"] = datos

    fecha = hoy - timedelta(days=dias["tiposdeatencion"])
    tdas = reportes.buscarTDA(practicas, fecha)
    habilitados = tdas.filter(baja=False)
    deshabilitados = tdas.filter(baja=True)
    tiposDeAtencion = {
        "fecha" : fecha,
        "dias" : dias["tiposdeatencion"],
        "habilitados" : habilitados.count(),
        "deshabilitados" : deshabilitados.count(),
    }
    if habilitados:
        normales, raros, descarte = reportes.clasificar(reportes.preparar(habilitados))
        datosNormales = reportes.datosTiposDeAtencion(normales)
        datosRaros = reportes.datosTiposDeAtencion(raros)
        tiposDeAtencion["normales"] = normales
        tiposDeAtencion["raros"] = raros
        tiposDeAtencion["descarte"] = descarte
        tiposDeAtencion["datos"] = {
            "normales" : datosNormales,
            "raros" : datosRaros,
        }

    return {
        "hoy" : hoy,
        "tipo" : area.nombre(),
        "perfiles" : perfiles,
        "realizacionesPorDia" : realizacionesPorDia,
        "porcentajesActualizacion" : porcentajesActualizacion,
        "tiposDeAtencion" : tiposDeAtencion,
    }


def realizaciones(request):

    usuario = request.user
    if isinstance(usuario, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (config("login_url"), request.path))

    puedeRealizar = permisos.filtroPermitidas(usuario, Practica.Acciones.realizar)
    puedeListar = permisos.filtroPermitidas(usuario, Practica.Acciones.listar)
    if (puedeListar is None) or (puedeRealizar is None):
        raise PermissionDenied()

    practicas = Estado.anotarPracticas(
        Practica.objects.enEstado([Realizada, Facturada]).realizadasPor(usuario),
        estado_actual=True,
        atendida_por=True
    ).filter(puedeListar)

    gestor = GestorListadoRealizacion(
        campos = [
            ["iniciorealizacion", "Inicio"],
            ["duracionrealizacion", "Duración"],
            ["id", "Práctica"],
            ["estadoactual", "Estado Actual"],
            ["cliente", "Cliente"],
            ["mascota", "Mascota"],
            ["tipodeatencion", "Tipo de Atención"],
            ["precio", "Precio"],
            ["marcacreacion", "Creación"],
            ["marcaultimaactualizacion", "Última Actualización"],
            ["creadapor", "Creada por"],
            ["atendidapor", "Atendida por"],
        ],
        iniciales = {
            "seleccion" : (
                "iniciorealizacion",
                "duracionrealizacion",
                "id",
                "estadoactual",
                "cliente",
            ),
            "orden" : (
                ("iniciorealizacion", False),
            ),
        },
        clases={"filtrado" : FiltradoRealizacionesForm},
        queryset=practicas,
        mapaFiltrado=practicas.MAPEO_FILTRADO,
        mapaOrden=practicas.MAPEO_ORDEN,
    )

    gestor.cargar(request)
    gestor.paginacion["cantidad"].label = "Practicas por página"

    menu = [[], []]
    itemListar(usuario, menu[0], Areas.C)
    itemListar(usuario, menu[0], Areas.Q)
    itemListarTurnos(usuario, menu[0], Areas.Q)
    itemCrear(usuario, menu[1], Areas.C)
    itemCrear(usuario, menu[1], Areas.Q)

    template = loader.get_template( plantilla("listar", "realizaciones") )
    context = {
        "tipo" : None,
        "gestor" : gestor,
        "practicas" : practicas,
        "menu" : [ seccion for seccion in menu if len(seccion) ],
        "visibles" : permisos.permitidas(
            usuario,
            Practica.Acciones.ver_informacion_general,
            gestor.itemsActuales()
        )
    }
    return HttpResponse(template.render( context, request ))
