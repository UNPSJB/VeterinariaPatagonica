from django.shortcuts import render
from django.template import loader
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from .models import Mascota
from .forms import MascotaFormFactory, FiltradoForm
from VeterinariaPatagonica import tools
from dal import autocomplete
from django.db.models import Q
from django.urls import reverse_lazy
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from Apps.GestionDeClientes.models import Cliente
from VeterinariaPatagonica.tools import GestorListadoQuerySet

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Importamos settings para poder tener a la mano la ruta de la carpeta media
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from django.views.generic import View
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import cm
from reportlab.lib import colors

mascotasFiltradas = []

def mascota(request):
    context = {}
    template = loader.get_template('GestionDeMascotas/GestionDeMascotas.html')
    return HttpResponse(template.render(context, request))


@permission_required('GestionDeMascotas.deshabilitar_mascota', raise_exception=True)
def dispatch(self, *args, **kwargs):
    return super(Mascota, self).dispatch(*args, **kwargs)

def menuVer(usuario, mascota):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeMascotas.mascota_modificar"):
        menu[0].append( (reverse("mascotas:mascotaModificar", args=(mascota.id,)), "Modificar mascota") )
        if mascota.baja:
            menu[0].append( (reverse("mascotas:mascotaHabilitar", args=(mascota.id,)), "Habilitar mascota") )
        else:
            menu[0].append( (reverse("mascotas:mascotaDeshabilitar", args=(mascota.id,)), "Deshabilitar mascota") )

    if usuario.has_perm("GestionDeMascotas.mascota_listar_habilitados"):
        menu[1].append( (reverse("mascotas:mascotaVerHabilitados"), "Listar mascotas habilitadas") )
    if usuario.has_perm("GestionDeMascotas.mascota_listar_no_habilitados"):
        menu[1].append( (reverse("mascotas:mascotaVerDeshabilitados"), "Listar mascotas deshabilitadas") )

    if usuario.has_perm("GestionDeMascotas.mascota_crear"):
        menu[2].append( (reverse("mascotas:mascotaCrear"), "Crear mascota") )

    return [ item for item in menu if len(item) ]

def menuListar(usuario, habilitados):
    menu = [[], [], [], [], []]

    if (not habilitados) and usuario.has_perm("GestionDeMascotas.mascota_ver_habilitados"):

        menu[0].append( (reverse("mascotas:mascotaVerHabilitados"), "Listar mascotas habilitadas") )
        menu[1].append( (reverse("mascotas:mascotasListadoExcel"), "Exportar mascotas deshabilitados"))
        menu[2].append( (reverse("mascotas:mascotasListadoPDF"), "Imprimir mascotas deshabilitados"))
    if habilitados and usuario.has_perm("GestionDeMascotas.mascota_ver_no_habilitados"):

        menu[0].append( (reverse("mascotas:mascotaVerDeshabilitados"), "Listar mascotas deshabilitadas") )
        menu[1].append( (reverse("mascotas:mascotasListadoExcel"), "Exportar mascotas habilitados") )
        menu[2].append( (reverse("mascotas:mascotasListadoPDF"), "Imprimir mascotas habilitados") )
    if usuario.has_perm("GestionDeMascotas.mascota_crear"):
        menu[3].append( (reverse("mascotas:mascotaCrear"), "Crear Mascota") )

    return [ item for item in menu if len(item) ]

def menuModificar(usuario, mascota):

    menu = [[],[],[],[],[]]

    menu[0].append( (reverse("mascotas:mascotaVer", args=(mascota.id,)), "Ver mascota") )

    if mascota.baja:
        menu[1].append( (reverse("mascotas:mascotaHabilitar", args=(mascota.id,)), "Habilitar mascota") )
    else:
        menu[1].append( (reverse("mascotas:mascotaDeshabilitar", args=(mascota.id,)), "Deshabilitar mascota") )

    if usuario.has_perm("GestionDeMascotas.mascota_listar_habilitados"):
        menu[2].append( (reverse("mascotas:mascotaVerHabilitados"), "Listar mascotas habilitadas") )
    if usuario.has_perm("GestionDeMascotas.mascota_listar_no_habilitados"):
        menu[2].append( (reverse("mascotas:mascotaVerDeshabilitados"), "Listar mascotas deshabilitadas") )

    if usuario.has_perm("GestionDeMascotas.mascota_crear"):
        menu[3].append( (reverse("mascotas:mascotaCrear"), "Crear mascota") )

    return [ item for item in menu if len(item) ]

def menuCrear(usuario, mascota):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeMascotas.mascota_listar_habilitados"):
        menu[0].append( (reverse("mascotas:mascotaVerHabilitados"), "Listar mascotas habilitados") )
    if usuario.has_perm("GestionDeMascotas.mascota_listar_no_habilitados"):
        menu[0].append( (reverse("mascotas:mascotaVerDeshabilitados"), "Listar mascotas deshabilitados") )

    return [ item for item in menu if len(item) ]

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeMascotas.add_Mascota', raise_exception=True)
def modificar(request, id= None, cliente_id=None):
    mascota = Mascota.objects.get(id=id) if id is not None else None
    cliente = None
    if cliente_id:
        cliente = Cliente.objects.get(pk=cliente_id)
    MascotaForm = MascotaFormFactory(mascota, cliente)

    if (id==None):
        context = {"titulo": 1, 'usuario': request.user}
    else:
        context = {"titulo": 2, 'usuario': request.user}

    if request.method == 'POST':
        formulario = MascotaForm(request.POST, instance=mascota)
        
        if formulario.is_valid():
            mascota = formulario.save()
            mascota.generadorDePatente(mascota.id)
            mascota = formulario.save()
            return HttpResponseRedirect("/GestionDeMascotas/ver/{}".format(mascota.id))
        else:
            context['form'] = formulario
            context['menu'] = menuModificar(request.user, mascota)
    else:
        context['form'] = MascotaForm(instance=mascota)
        context['menu'] = menuCrear(request.user, mascota)
    template = loader.get_template('GestionDeMascotas/formulario.html')
    return HttpResponse(template.render(context, request))

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeMascotas.delete_Mascota', raise_exception=True)
def habilitar(request, id):

    try:
        mascota= Mascota.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    mascota.baja = False
    mascota.save()

    return HttpResponseRedirect( "/GestionDeMascotas/verHabilitados/" )


@login_required(redirect_field_name='proxima')
@permission_required('GestionDeMascotas.delete_Mascotas', raise_exception=True)
def deshabilitar(request, id):
    try:
        mascota = Mascota.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    mascota.baja = True
    mascota.save()

    return HttpResponseRedirect( "/GestionDeMascotas/verDeshabilitados/" )

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeMascotas.delete_Mascota', raise_exception=True)
def eliminar(request, id):
    try:
        mascota = Mascota.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()
    if request.method == 'POST':
        mascota.delete()
        return HttpResponseRedirect( "/GestionDeMascotas/verDeshabilitados/" )
    else:
        template = loader.get_template('GestionDeMascotas/eliminar.html')
        context = {
            'usuario' : request.user,
            'id' : id
        }
        return HttpResponse( template.render( context, request) )

def ver(request, id):

    try:
        mascota = Mascota.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404("No encontrado", "La mascota con patente={} no existe.".format(id))


    template = loader.get_template('GestionDeMascotas/ver.html')
    contexto = {
        'mascota': mascota,
        'usuario': request.user,
        "menu" : menuVer(request.user, mascota)
    }
    return HttpResponse(template.render(contexto, request))



def verHabilitados(request, habilitados=True):
    
    global mascotasFiltradas
    mascotas = Mascota.objects.habilitados()
    gestor = GestorListadoQuerySet(
        campos=[
            ["orden_patente", "Patente"],
            ["orden_nombre", "Nombre"],
            ["orden_cliente", "Dueño"],
            ["orden_especie", "Especie"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=mascotas,
        mapaFiltrado= Mascota.MAPPER,
        mapaOrden= mascotas.MAPEO_ORDEN
    )
  
    gestor.cargar(request)
    mascotasFiltradas = gestor.queryset
    template = loader.get_template('GestionDeMascotas/verHabilitados.html')
    contexto = {
        "gestor" : gestor,
        "menu" : menuListar(request.user, habilitados),}
    return HttpResponse(template.render(contexto, request))

def verDeshabilitados(request, habilitados=False):
    
    global mascotasFiltradas
    mascotas = Mascota.objects.deshabilitados()
    gestor = GestorListadoQuerySet(
        campos=[
            ["orden_patente", "Patente"],
            ["orden_nombre", "Nombre"],
            ["orden_cliente", "Dueño"],
            ["orden_especie", "Especie"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=mascotas,
        mapaFiltrado= Mascota.MAPPER,
        mapaOrden= mascotas.MAPEO_ORDEN
    )
  
    gestor.cargar(request)
    mascotasFiltradas = gestor.queryset
    template = loader.get_template('GestionDeMascotas/verDeshabilitados.html')
    contexto = {
        "gestor" : gestor,
        "menu" : menuListar(request.user, habilitados),}
    return HttpResponse(template.render(contexto, request))

def ListadoMascotasExcel(request):
    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'LISTADO DE MASCOTAS'
    ws.merge_cells('B1:E1')
    ws['B3'] = 'PATENTE'
    ws['C3'] = 'NOMBRE'
    ws['D3'] = 'DUEÑO'
    ws['E3'] = 'ESPECIE'
    cont = 4
    for mascota in mascotasFiltradas:
        ws.cell(row=cont, column=2).value = mascota.patente
        ws.cell(row=cont, column=3).value = mascota.nombre
        ws.cell(row=cont, column=4).value = str(mascota.cliente)
        ws.cell(row=cont, column=5).value = mascota.especie
        cont = cont + 1
    
    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]

    for i, column_width in enumerate(column_widths):
         ws.column_dimensions[get_column_letter(i+1)].width = column_width


    nombre_archivo = "ListadoMascotas.xlsx"

    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

def ListadoMascotasPDF(request):
    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)
    cabecera(pdf)
    y = 650
    tabla(pdf, y, mascotasFiltradas)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def cabecera(pdf):
    archivo_imagen = settings.MEDIA_ROOT + '/imagenes/logo_vetpat2.jpeg'
    pdf.drawImage(archivo_imagen, 20, 750, 120, 90, preserveAspectRatio=True)
    pdf.setFont("Helvetica", 16)
    pdf.drawString(190, 790, u"VETERINARIA PATAGONICA")
    pdf.setFont("Helvetica", 14)
    pdf.drawString(220, 770, u"LISTADO DE MASCOTAS")

def tabla(pdf, y, mascotas):
    encabezados = ('Patente', 'Nombre', 'Dueño', 'Especie')

    detalles = [(mascota.patente, mascota.nombre, mascota.cliente, mascota.especie) for mascota in mascotas]
    detalle_orden = Table([encabezados] + detalles, colWidths=[3 * cm, 4 * cm, 5 * cm, 4 * cm])
    detalle_orden.setStyle(TableStyle(
        [
            ('ALIGN', (0, 0), (3, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]
    ))
    detalle_orden.wrapOn(pdf, 800, 600)
    detalle_orden.drawOn(pdf, 65, y)

@login_required
def ayudaContextualMascota(request):

    template = loader.get_template('GestionDeMascotas/AyudaGestiondeMascotas.html')
    contexto = {
        'usuario': request.user,
    }

    return HttpResponse(template.render(contexto, request))


class clienteAutocomplete(autocomplete.Select2QuerySetView):

    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !

        qs = Cliente.objects.all()

        if self.q:
            qs = qs.filter(Q(apellidos__icontains=self.q) |Q(nombres__icontains=self.q) | Q(dniCuit__icontains=self.q))

        return qs

