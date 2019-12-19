from django.template import loader
from django.shortcuts import render
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ObjectDoesNotExist
from .models import Producto
from .forms import ProductoFormFactory, FiltradoForm
from VeterinariaPatagonica import tools
from dal import autocomplete
from django.db.models import Q
from django.urls import reverse_lazy
from django.urls import reverse
from Apps.GestionDeRubros.models import Rubro
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from VeterinariaPatagonica.tools import GestorListadoQuerySet
from VeterinariaPatagonica.forms import ExportarForm
from Apps.GestionDePracticas.models import *

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Importamos settings para poder tener a la mano la ruta de la carpeta media
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from django.views.generic import View
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import cm
from reportlab.lib import colors

productosFiltrados = []

def producto(request):
    context = {}#Defino un contexto.
    template = loader.get_template('GestionDeProductos/GestionDeProductos.html')#Cargo el template desde la carpeta templates/GestionDeProductos.
    return HttpResponse(template.render(context, request))#Devuelvo la url con el template armado.

@permission_required('GestionDeProductos.deshabilitar_producto', raise_exception=True)
def dispatch(self, *args, **kwargs):
    return super(Producto, self).dispatch(*args, **kwargs)

def menuVer(usuario, producto):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeProductos.producto_modificar"):
        menu[0].append( (reverse("productos:productoModificar", args=(producto.id,)), "Modificar producto/insumo") )
        if producto.baja:
            menu[0].append( (reverse("productos:productoHabilitar", args=(producto.id,)), "Habilitar producto/insumo") )
        else:
            menu[0].append( (reverse("productos:productoDeshabilitar", args=(producto.id,)), "Deshabilitar producto/insumo") )

    if usuario.has_perm("GestionDeProductos.producto_listar_habilitados"):
        menu[1].append( (reverse("productos:productoVerHabilitados"), "Listar productos/insumos habilitados") )
    if usuario.has_perm("GestionDeProductos.producto_listar_no_habilitados"):
        menu[1].append( (reverse("productos:productoVerDeshabilitados"), "Listar productos/insumos deshabilitados") )

    if usuario.has_perm("GestionDeProductos.producto_crear"):
        menu[2].append( (reverse("productos:productoCrear"), "Crear producto/insumo") )



    return [ item for item in menu if len(item) ]

def menuListar(usuario, habilitados):
    menu = [[], [], [], [], []]

    if (not habilitados) and usuario.has_perm("GestionDeProductos.producto_ver_habilitados"):

        menu[0].append( (reverse("productos:productoVerHabilitados"), "Listar productos/insumos habilitados") )
        menu[1].append( (reverse("productos:productosListadoExcel"), "Exportar productos/insumos deshabilitados"))
        menu[2].append( (reverse("productos:productosListadoPDF"), "Imprimir productos/insumos deshabilitados"))

    if habilitados and usuario.has_perm("GestionDeProductos.producto_ver_no_habilitados"):
        menu[0].append( (reverse("productos:productoVerDeshabilitados"), "Listar productos/insumos deshabilitados") )
        menu[1].append( (reverse("productos:productosListadoExcel"), "Exportar productos/insumos habilitados") )
        menu[2].append( (reverse("productos:productosListadoPDF"), "Imprimir productos/insumos habilitados") )

    if usuario.has_perm("GestionDeProductos.producto_crear"):
        menu[3].append( (reverse("productos:productoCrear"), "Crear Producto/Insumo") )


    return [ item for item in menu if len(item) ]

def menuModificar(usuario, producto):

    menu = [[],[],[],[],[]]

    menu[0].append( (reverse("productos:productoVer", args=(producto.id,)), "Ver producto/insumo") )

    if producto.baja:
        menu[1].append( (reverse("productos:productoHabilitar", args=(producto.id,)), "Habilitar producto/insumo") )
    else:
        menu[1].append( (reverse("productos:productoDeshabilitar", args=(producto.id,)), "Deshabilitar producto/insumo") )

    if usuario.has_perm("GestionDeProductos.producto_listar_habilitados"):
        menu[2].append( (reverse("productos:productoVerHabilitados"), "Listar productos/insumos habilitados") )
    if usuario.has_perm("GestionDeProductos.producto_listar_no_habilitados"):
        menu[2].append( (reverse("productos:productoVerDeshabilitados"), "Listar productos/insumos deshabilitados") )

    if usuario.has_perm("GestionDeProductos.producto_crear"):
        menu[3].append( (reverse("productos:productoCrear"), "Crear producto/insumo") )


    return [ item for item in menu if len(item) ]

def menuCrear(usuario, producto):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeProductos.producto_listar_habilitados"):
        menu[0].append( (reverse("productos:productoVerHabilitados"), "Listar productos/insumos habilitados") )
    if usuario.has_perm("GestionDeProductos.producto_listar_no_habilitados"):
        menu[0].append( (reverse("productos:productoVerDeshabilitados"), "Listar productos/insumos deshabilitados") )
    return [ item for item in menu if len(item) ]


@login_required(redirect_field_name='proxima')
@permission_required('GestionDeProductos.add_Producto', raise_exception=True)
def modificar(request, id = None, producto_id=None):
    producto = Producto.objects.get(id=id) if id is not None else None
    ProductoForm = ProductoFormFactory(producto)

    if (id==None):
        context = {"titulo": 1, 'usuario': request.user}
    else:
        context = {"titulo": 2, 'usuario': request.user}

    if request.method == 'POST':
        formulario = ProductoForm(request.POST, instance=producto)
        if formulario.is_valid():
            producto = formulario.save()
            return HttpResponseRedirect("/GestionDeProductos/ver/{}".format(producto.id))
        else:
            context['form'] = formulario
            context['menu'] = menuModificar(request.user, producto)
    else:
        context['form'] = ProductoForm(instance=producto)
        context['menu'] = menuCrear(request.user, producto)
    template = loader.get_template('GestionDeProductos/formulario.html')
    return HttpResponse(template.render(context, request))

def verHabilitados(request, habilitados=True):
    global productosFiltrados
    productos = Producto.objects.habilitados()
    gestor = GestorListadoQuerySet(
        campos=[
            ["orden_nombre", "Nombre"],
            ["orden_marca", "Marca"],
            ["orden_formaDePresentacion", "Forma de Presentacion"],
            ["orden_precioPorUnidad", "Precio por Unidad"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=productos,
        mapaFiltrado= Producto.MAPPER,
        mapaOrden= productos.MAPEO_ORDEN
    )

    gestor.cargar(request)
    productosFiltrados = gestor.queryset
    template = loader.get_template('GestionDeProductos/verHabilitados.html')
    contexto = {
        "gestor" : gestor,
        "menu" : menuListar(request.user, habilitados)
    }
    return HttpResponse(template.render(contexto, request))

def verDeshabilitados(request, habilitados=False):
    global productosFiltrados
    productos = Producto.objects.deshabilitados()
    gestor = GestorListadoQuerySet(
        campos=[
            ["orden_nombre", "Nombre"],
            ["orden_marca", "Marca"],
            ["orden_formaDePresentacion", "Forma de Presentacion"],
            ["orden_precioPorUnidad", "Precio por Unidad"],
        ],
        clases={"filtrado" : FiltradoForm},
        queryset=productos,
        mapaFiltrado= Producto.MAPPER,
        mapaOrden= productos.MAPEO_ORDEN
    )

    gestor.cargar(request)
    productosFiltrados = gestor.queryset

    template = loader.get_template('GestionDeProductos/verDeshabilitados.html')
    contexto = {
        "gestor" : gestor,
        "menu" : menuListar(request.user, habilitados)
    }
    return HttpResponse(template.render(contexto, request))

def ver(request, id):
    try:
        producto = Producto.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404( "No encontrado", "El Producto con id={} no existe.".format(id))

    template = loader.get_template('GestionDeProductos/ver.html')
    context = {
        'producto': producto,
        'usuario': request.user,
        'menu' : menuVer(request.user, producto)
    }
    return HttpResponse(template.render(context, request))

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeProductos.delete_Producto', raise_exception=True)
def deshabilitar(request, id):

    try:
        producto = Producto.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    '''practicas = Practica.objects.enEstado([Presupuestada]).filter(producto=producto).count()

    if practicas > 0:
        raise VeterinariaPatagonicaError("Error","El producto se encuentra en practicas presupuestadas")'''
    

    producto.baja = True
    producto.save()

    return HttpResponseRedirect( "/GestionDeProductos/verDeshabilitados/" )

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeProductos.delete_Producto', raise_exception=True)
def habilitar(request, id):
    try:
        producto = Producto.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    producto.baja = False
    producto.save()

    return HttpResponseRedirect( "/GestionDeProductos/verHabilitados/" )


@login_required(redirect_field_name='proxima')
@permission_required('GestionDeProductos.delete_Producto', raise_exception=True)
def eliminar(request, id):

    try:
        producto = Producto.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()
    if request.method == 'POST':
        producto.delete()
        return HttpResponseRedirect( "/GestionDeProductos/verDeshabilitados/" )
    else:
        template = loader.get_template('GestionDeProductos/eliminar.html')
        context = {
            'usuario' : request.user,
            'id' : id
        }
        return HttpResponse( template.render( context, request) )

class rubroAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor!
        qs = Rubro.objects.all()
        if self.q:
            qs = qs.filter(Q(nombre__icontains=self.q))

        return qs


@login_required
def ayudaContextualProducto(request):

    template = loader.get_template('GestionDeProductos/ayudaContextualProducto.html')
    contexto = {
        'usuario': request.user,
    }
    return HttpResponse(template.render(contexto, request))

def ListadoProductosExcel(request):
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'LISTADO DE PRODUCTOS'
    ws['B1'] = 'LISTADO DE PRODUCTOS'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:F1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'MARCA'
    ws['D3'] = 'FORMA DE PRESENTACIÓN'
    ws['E3'] = 'PRECIO POR UNIDAD'
    cont = 4
    # Recorremos el conjunto de productos y vamos escribiendo cada uno de los datos en las celdas
    for producto in productosFiltrados:
        ws.cell(row=cont, column=2).value = producto.nombre
        ws.cell(row=cont, column=3).value = producto.marca
        ws.cell(row=cont, column=4).value = producto.formaDePresentacion
        ws.cell(row=cont, column=5).value = producto.precioPorUnidad
        cont = cont + 1

    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]

    for i, column_width in enumerate(column_widths):
         ws.column_dimensions[get_column_letter(i+1)].width = column_width

    # Establecemos el nombre del archivo
    nombre_archivo = "ListadoProductos.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

def ListadoProductosPDF(request):
    print("GET")
    # Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    # La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    # Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = canvas.Canvas(buffer)
    # Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera(pdf)
    y = 730
    tabla(pdf, y, productosFiltrados)
    # Con show page hacemos un corte de página para pasar a la siguiente
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def cabecera(pdf):
    print("CABECERA")
    # Utilizamos el archivo logo_vetpat.png que está guardado en la carpeta media/imagenes
    archivo_imagen = settings.MEDIA_ROOT + '/imagenes/logo_vetpat2.jpeg'
    # Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 20, 750, 120, 90, preserveAspectRatio=True)
    # Establecemos el tamaño de letra en 16 y el tipo de letra Helvetica
    pdf.setFont("Helvetica", 16)
    # Dibujamos una cadena en la ubicación X,Y especificada
    pdf.drawString(190, 790, u"VETERINARIA PATAGÓNICA")
    pdf.setFont("Helvetica", 14)
    pdf.drawString(220, 770, u"LISTADO DE PRODUCTOS")

def tabla(pdf, y, productos):
    print("TABLA")
    # Creamos una tupla de encabezados para neustra tabla
    encabezados = ('Nombre', 'Marca', 'Forma de Presentación', 'Precio')
    # Creamos una lista de tuplas que van a contener a los productos
    detalles = []
    for producto in productos:
        y -= 20
        p = (producto.nombre, producto.marca, producto.formaDePresentacion, producto.precioPorUnidad)
        detalles.append(p)

    # Establecemos el tamaño de cada una de las columnas de la tabla
    detalle_orden = Table([encabezados] + detalles, colWidths=[5 * cm, 4 * cm, 4 * cm, 4 * cm])
    # Aplicamos estilos a las celdas de la tabla
    detalle_orden.setStyle(TableStyle(
        [
            # La primera fila(encabezados) va a estar centrada
            ('ALIGN', (0, 0), (3, 0), 'CENTER'),
            # Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            # El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]
    ))
    # Establecemos el tamaño de la hoja que ocupará la tabla
    detalle_orden.wrapOn(pdf, 800, 600)
    # Definimos la coordenada donde se dibujará la tabla
    detalle_orden.drawOn(pdf, 60, y)

class rubroAutocomplete(autocomplete.Select2QuerySetView):

    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        qs = Rubro.objects.all()
        if self.q:
            qs = qs.filter(Q(nombre__icontains=self.q))

        return qs
