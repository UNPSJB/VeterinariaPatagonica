from django.shortcuts import render
from django.template import loader
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required, permission_required
from django.forms import modelformset_factory

from .models import Servicio, ServicioProducto
from .forms import ServicioForm, ServicioProductoForm, ServicioProductoBaseFormSet
from VeterinariaPatagonica import tools

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Importamos settings para poder tener a la mano la ruta de la carpeta media
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from django.views.generic import View
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import cm
from reportlab.lib import colors


serviciosFiltrados = []

def menuVer(usuario, servicio):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeServicios.servicio_modificar"):
        menu[0].append( (reverse("servicios:servicioModificar", args=(servicio.id,)), "Modificar servicio") )
        if servicio.baja:
            menu[0].append( (reverse("servicios:servicioHabilitar", args=(servicio.id,)), "Habilitar servicio") )
        else:
            menu[0].append( (reverse("servicios:servicioDeshabilitar", args=(servicio.id,)), "Deshabilitar servicio") )

    if usuario.has_perm("GestionDeServicios.servicio_listar_habilitados"):
        menu[1].append( (reverse("servicios:servicioVerHabilitados"), "Listar servicios habilitados") )
    if usuario.has_perm("GestionDeServicios.servicio_listar_no_habilitados"):
        menu[1].append( (reverse("servicios:servicioVerDeshabilitados"), "Listar servicios deshabilitados") )

    if usuario.has_perm("GestionDeServicios.servicio_crear"):
        menu[2].append( (reverse("servicios:servicioCrear"), "Crear servicio") )

    return [ item for item in menu if len(item) ]

def menuListar(usuario, habilitados):
    menu = [[],[],[],[],[]]

    if (not habilitados) and usuario.has_perm("GestionDeServicios.cliente_ver_habilitados") \
        and usuario.has_perm("GestionDeServicios.cliente_exportar_excel_habilitados"):

        menu[0].append( (reverse("servicios:servicioVerHabilitados"), "Listar servicios habilitados") )
        menu[1].append( (reverse("servicios:serviciosListadoEXCEL"), "Exportar servicios deshabilitados"))
        menu[2].append( (reverse("servicios:serviciosListadoPDF"), "Imprimir servicios deshabilitados"))

    if habilitados and usuario.has_perm("GestionDeServicios.cliente_ver_no_habilitados") \
        and usuario.has_perm("GestionDeServicios.cliente_exportar_excel_deshabilitados"):

        menu[0].append( (reverse("servicios:servicioVerDeshabilitados"), "Listar servicios deshabilitados") )
        menu[1].append( (reverse("servicios:serviciosListadoEXCEL"), "Exportar servicios habilitados") )
        menu[2].append( (reverse("servicios:serviciosListadoPDF"), "Imprimir servicios habilitados") )

    if usuario.has_perm("GestionDeServicios.servicio_crear"):

        menu[3].append( (reverse("servicios:servicioCrear"), "Crear Servicio") )

    return [ item for item in menu if len(item) ]

def menuModificar(usuario, servicio):

    menu = [[],[],[],[],[]]
    if (servicio):
        menu[0].append( (reverse("servicios:servicioVer", args=(servicio.id,)), "Ver servicio") )

        if servicio.baja:
            menu[1].append( (reverse("servicios:servicioHabilitar", args=(servicio.id,)), "Habilitar servicio") )
        else:
            menu[1].append( (reverse("servicios:servicioDeshabilitar", args=(servicio.id,)), "Deshabilitar servicio") )

    if usuario.has_perm("GestionDeServicios.servicio_listar_habilitados"):
        menu[2].append( (reverse("servicios:servicioVerHabilitados"), "Listar servicios habilitados") )
    if usuario.has_perm("GestionDeServicios.servicio_listar_no_habilitados"):
        menu[2].append( (reverse("servicios:servicioVerDeshabilitados"), "Listar servicios deshabilitados") )

    if usuario.has_perm("GestionDeServicios.servicio_crear"):
        menu[3].append( (reverse("servicios:servicioCrear"), "Crear servicio") )

    return [ item for item in menu if len(item) ]

def menuCrear(usuario, servicio):

    menu = [[],[],[],[]]

    if usuario.has_perm("GestionDeServicios.servicio_listar_habilitados"):
        menu[0].append( (reverse("servicios:servicioVerHabilitados"), "Listar servicios habilitados") )
    if usuario.has_perm("GestionDeServicios.servicio_listar_no_habilitados"):
        menu[0].append( (reverse("servicios:servicioVerDeshabilitados"), "Listar servicios deshabilitados") )

    return [ item for item in menu if len(item) ]


@login_required(redirect_field_name='proxima')
@permission_required('GestionDeServicios.add_Servicio', raise_exception=True)
def modificar(request, id = None):
    servicio = Servicio.objects.get(id=id) if id is not None else None
    if (id==None):
        context = {"titulo": 1, 'usuario': request.user}
    else:
        context = {"titulo": 2, 'usuario': request.user}
    form = ServicioForm(instance=servicio)
    ServicioProductoFormset = modelformset_factory(ServicioProducto,#          Defino la forma del formset. Van a tener el checkbox eliminar, la cantidad mínima de forms en el formset
        fields=("producto", "cantidad"), min_num=1, extra=0, can_delete=True,# es de 1 y la máxima es la determinada por django (1000),
        formset=ServicioProductoBaseFormSet)#                                  además se define que siempre va a haber cero tuplas adicionales (extra).
    if request.method == 'POST':
        form = ServicioForm(request.POST, instance=servicio)
        formset = ServicioProductoFormset(request.POST)
        if form.is_valid() and formset.is_valid():#Verifico si los forms del formset cumplen las restricciones definidas (Si no se lanzó un ValidationError).
            servicio = form.save()
            instances = formset.save(commit=False)
            for obj in formset.deleted_objects:#Bucle for que elimina los form que tienen tildado el checkbox "eliminar"
                obj.delete()
            for sproducto in instances:
                sproducto.servicio = servicio
                sproducto.save()

            return HttpResponseRedirect("/GestionDeServicios/ver/{}".format(servicio.id))
        context['form'] = form
        context["formset"] = formset
        context['menu'] = menuModificar(request.user, servicio)
    else:
        context['form'] = form
        qs = ServicioProducto.objects.none() if servicio is None else servicio.servicio_productos.all()#Obtengo los productos ya cargados en el servicio, para poder mostrarlos en el formset.
        context["formset"] = ServicioProductoFormset(queryset=qs)
        context['menu'] = menuCrear(request.user, servicio)
    template = loader.get_template('GestionDeServicios/formulario.html')
    return HttpResponse(template.render(context, request))


# def verHabilitados(request):
#     serviciosQuery = Servicio.objects.habilitados()
#     serviciosQuery = serviciosQuery.filter(tools.paramsToFilter(request.GET, Servicio))
#     template = loader.get_template('GestionDeServicios/verHabilitados.html')

#     paginator = Paginator(serviciosQuery, 1)
#     page = request.GET.get('page')

#     try:
#         servicios = paginator.page(page)
#     except PageNotAnInteger:
#         # If page is not an integer, deliver first page.
#         servicios = paginator.page(1)
#     except EmptyPage:
#         # If page is out of range (e.g. 9999), deliver last page of results.
#         servicios = paginator.page(paginator.num_pages)

#     context = {
#         'serviciosQuery' : serviciosQuery,
#         'usuario' : request.user,
#         'servicios': servicios,
#     }

#     return  HttpResponse(template.render(context,request))

# def verDeshabilitados(request):
#     serviciosQuery = Servicio.objects.deshabilitados()
#     serviciosQuery = serviciosQuery.filter(tools.paramsToFilter(request.GET, Servicio))
#     template = loader.get_template('GestionDeServicios/verDeshabilitados.html')

#     paginator = Paginator(serviciosQuery, 1)
#     page = request.GET.get('page')

#     try:
#         servicios = paginator.page(page)
#     except PageNotAnInteger:
#         # If page is not an integer, deliver first page.
#         servicios = paginator.page(1)
#     except EmptyPage:
#         # If page is out of range (e.g. 9999), deliver last page of results.
#         servicios = paginator.page(paginator.num_pages)

#     context = {
#         'serviciosQuery': serviciosQuery,
#         'usuario': request.user,
#         'servicios': servicios,
#     }

#     return  HttpResponse(template.render(context,request))

def ver(request, id):
    #import ipdb; ipdb.set_trace()
    try:
        servicio = Servicio.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404( "No encontrado", "El servicio con id={} no existe.".format(id) )
    template = loader.get_template('GestionDeServicios/ver.html')
    context = {
        'servicio' : servicio,
        'usuario' : request.user,
        'menu': menuVer(request.user, servicio)
    }
    return HttpResponse(template.render( context, request ))

@login_required(redirect_field_name='proxima')
@permission_required('GestionDeServicios.delete_Servicio', raise_exception=True)
def deshabilitar(request, id):

    try:
        servicio = Servicio.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    servicio.baja = True
    servicio.save()

    return HttpResponseRedirect( "/GestionDeServicios/verDeshabilitados/")


@login_required(redirect_field_name='proxima')
@permission_required('GestionDeServicios.delete_Servicio', raise_exception=True)
def habilitar(request, id):

    try:
        servicio = Servicio.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()

    servicio.baja = False
    servicio.save()

    return HttpResponseRedirect( "/GestionDeServicios/verHabilitados/")


@login_required(redirect_field_name='proxima')
@permission_required('GestionDeServicios.delete_Servicio', raise_exception=True)
def eliminar(request, id):
    try:
        servicio = Servicio.objects.get(id=id)
    except ObjectDoesNotExist:
        raise Http404()
    if request.method == 'POST':
        servicio.delete()
        return HttpResponseRedirect( "/GestionDeServicios/verDeshabilitados" )
    else:
        template = loader.get_template('GestionDeServicios/eliminar.html')
        context = {
            'usuario' : request.user,
            'id' : id
        }
        return HttpResponse( template.render( context, request) )

@login_required
def ayudaContextualServicio(request):

    template = loader.get_template('GestionDeServicios/ayudaContextualServicios.html')
    contexto = {
        'usuario': request.user,
    }

    return HttpResponse(template.render(contexto, request))
#def volver(request, id):
#    return HttpResponseRedirect(request.META.get('HTTP_REFERER','/'))


from django import forms
from django.contrib.auth.models import AnonymousUser
from django.core.exceptions import PermissionDenied
from VeterinariaPatagonica.tools import GestorListadoQuerySet
LOGIN_URL="/login/"


def listar(request, habilitados=True):

    global serviciosFiltrados

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    if not request.user.has_perm("GestionDeServicios.servicio_listar_habilitados"):
        raise PermissionDenied()

    if habilitados:
        queryset = Servicio.objects.habilitados()
    else:
        queryset = Servicio.objects.deshabilitados()

    gestor = GestorListadoQuerySet(
        campos = [
            ["nombre", "Nombre"],
            ["tipo", "Tipo"],
            ["precio", "Precio de mano de obra"],
        ],
        clases={"filtrado" : FiltradoServiciosForm},
        queryset=queryset,
        mapaFiltrado={
            "nombre": "nombre__icontains",
            "tipo": "tipo__icontains",
            "precio": "precioManoDeObra__icontains",
        },
        mapaOrden={
            "nombre" : ["nombre"],
            "tipo" : ["tipo"],
            "precio" : ["precioManoDeObra"],
        },
    )

    gestor.cargar(request)
    gestor.paginacion["cantidad"].label = "Servicios por página"

    serviciosFiltrados = gestor.queryset
    template = loader.get_template( "GestionDeServicios/listar.html" )
    context = {
        "gestor" : gestor,
        "tipo" : "habilitados" if habilitados else "deshabilitados",
        "menu" : menuListar(request.user, habilitados),
    }
    return HttpResponse(template.render( context, request ))

def ListadoServiciosExcel(request):
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    ws = wb.worksheets[0]
    # En la celda B1 ponemos el texto 'LISTADO DE SERVICIOS'
    ws['B1'] = 'LISTADO DE SERVICIOS'
    # Juntamos las celdas desde la B1 hasta la H1, formando una sola celda
    ws.merge_cells('B1:H1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'TIPO'
    ws['D3'] = 'PRECIO DE MANO DE OBRA'
    cont = 5
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas

    for servicio in serviciosFiltrados:
        ws.cell(row=cont, column=2).value = servicio.nombre
        ws.cell(row=cont, column=3).value = servicio.tipo
        ws.cell(row=cont, column=4).value = servicio.precioManoDeObra
        cont = cont + 1


    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]

    for i, column_width in enumerate(column_widths):
         ws.column_dimensions[get_column_letter(i+1)].width = column_width

    # Establecemos el nombre del archivo
    nombre_archivo = "ListadoServicios.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

def ListadoServiciosPDF(request):
    print("GET")
    # Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    # La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    # Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = canvas.Canvas(buffer)
    # Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera(pdf)
    y = 500
    tabla(pdf, y, serviciosFiltrados)
    # Con show page hacemos un corte de página para pasar a la siguiente
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def cabecera(pdf):
    print("CABECERA")
    # Utilizamos el archivo logo_vetpat.png que está guardado en la carpeta media/imagenes
    archivo_imagen = settings.MEDIA_ROOT + '/imagenes/logo_vetpat2.jpeg'
    # Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 20, 750, 120, 90, preserveAspectRatio=True)
    # Establecemos el tamaño de letra en 16 y el tipo de letra Helvetica
    pdf.setFont("Helvetica", 16)
    # Dibujamos una cadena en la ubicación X,Y especificada
    pdf.drawString(190, 790, u"VETERINARIA PATAGONICA")
    pdf.setFont("Helvetica", 14)
    pdf.drawString(220, 770, u"LISTADO DE SERVICIOS")

def tabla(pdf, y, servicios):
    print("TABLA")
    # Creamos una tupla de encabezados para neustra tabla
    encabezados = ('Nombre', 'Tipo', 'Precio de mano de obra')

    # Creamos una lista de tuplas que van a contener a las personas
    detalles = [(servicio.nombre, servicio.tipo, servicio.precioManoDeObra) for servicio in servicios]

    # Establecemos el tamaño de cada una de las columnas de la tabla
    detalle_orden = Table([encabezados] + detalles, colWidths=[5 * cm, 5 * cm, 5 * cm])
    print("HOLAAA")
    # Aplicamos estilos a las celdas de la tabla
    detalle_orden.setStyle(TableStyle(
        [
            # La primera fila(encabezados) va a estar centrada
            ('ALIGN', (0, 0), (3, 0), 'CENTER'),
            # Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            # El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]
    ))
    # Establecemos el tamaño de la hoja que ocupará la tabla
    detalle_orden.wrapOn(pdf, 800, 600)
    # Definimos la coordenada donde se dibujará la tabla
    detalle_orden.drawOn(pdf, 30, y)

class FiltradoServiciosForm(forms.Form):
    nombre = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={"class":"form-control", "placeholder" : "Nombre"})
    )
    tipo = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={"class":"form-control", "placeholder" : "Tipo"})
    )
    precio = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={"class":"form-control", "placeholder" : "Precio de mano de obra"})
    )
    def filtros(self):
        retorno={}
        if self.is_valid():
            retorno.update({
                campo:valor for campo,valor in self.cleaned_data.items() if valor
            })
        return retorno
