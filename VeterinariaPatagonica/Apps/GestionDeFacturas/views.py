from datetime import datetime

from django.template import loader
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib.auth.models import AnonymousUser
from django.core.exceptions import PermissionDenied
from django.core.validators import MinValueValidator
from django.db import transaction
from django.db.models import Sum, Q
from django.db.utils import Error as dbError

from VeterinariaPatagonica.forms import ExportarForm
from VeterinariaPatagonica.tools import GestorListadoQuerySet, guardarWorkbook
from VeterinariaPatagonica.errores import VeterinariaPatagonicaError
from Apps.GestionDePracticas import permisos
from Apps.GestionDeClientes.models import Cliente
from Apps.GestionDePagos.models import Pago
from Apps.GestionDePracticas.models.practica import Practica
from Apps.GestionDePracticas.gestionDePracticas import pathVer
from .models import Factura, DetalleFactura
from .forms import *

from openpyxl import Workbook

LOGIN_URL="/login/"



class GestorListadoFacturas(GestorListadoQuerySet):

    def hacerSeleccion(self, *args, **kwargs):
        super().hacerSeleccion(*args, **kwargs)

        if self.campos["importeVentas"]["visible"]:
            self.queryset = self.queryset.annotate(
                importe_ventas=Sum("detalles_producto__subtotal")
            )

def verificarFacturacion(usuario, practica):

    if not permisos.paraPractica(usuario, Practica.Acciones.facturar, practica):
        raise VeterinariaPatagonicaError(
            "Permiso denegado",
            "El usuario %s no tiene permisos suficientes para facturar la practica solicitada (%s)" % (usuario.username, str(practica))
        )

    if not practica.esPosible(Practica.Acciones.facturar):
        raise VeterinariaPatagonicaError(
            "%s no facturable" % practica.nombreTipo().capitalize(),
            "La %s no puede facturarse" % practica.nombreTipo()
        )

    if Factura.objects.filter(practica=practica).count():
        raise VeterinariaPatagonicaError(
            "%s facturada" % practica.nombreTipo().capitalize(),
            "No puede volver a facturarse la %s" % practica.nombreTipo()
        )

def menuCrear(usuario, practica=None):
    menu = [[], []]
    poseidos = usuario.get_all_permissions()

    if practica is not None:
        if permisos.paraPractica(usuario, Practica.Acciones.ver_informacion_general, practica):
            menu[0].append( (reverse(
                "practicas:%s:ver:practica" % practica.nombreTipo(), args=(practica.id,)
            ), "Ver practica") )

        cliente = practica.cliente if not practica.cliente.baja else None
        if cliente and "GestionDeClientes.cliente_ver_habilitados" in poseidos:
            menu[0].append( (reverse("clientes:clienteVer", args=(cliente.id,)), "Ver cliente") )

    permitidas = filtrosPara(usuario, "listar")
    if permitidas is not None:
        menu[1].append(
            (reverse("facturas:listar"), "Listar facturas")
        )
    if "GestionDePagos.pago_listar" in poseidos:
        menu[1].append( (reverse("pagos:listar"), "Listar pagos") )
    return [ item for item in menu if len(item)]

def menuVer(usuario, factura):
    menu = [[], [], []]
    poseidos = usuario.get_all_permissions()

    if factura.estaPaga():
        if "GestionDePagos.pago_ver" in poseidos:
            menu[1].append( (reverse("pagos:ver", args=(factura.pago.id,)), "Ver pago") )
    else:
        if "GestionDePagos.pago_crear" in poseidos:
            menu[0].append( (reverse("pagos:crear", args=(factura.id,)), "Pagar factura") )

    practica = factura.practica
    if practica is not None:
        if permisos.paraPractica(usuario, Practica.Acciones.ver_informacion_general, practica):
            menu[1].append( (reverse(
                "practicas:%s:ver:practica" % practica.nombreTipo(), args=(practica.id,)
            ), "Ver practica") )

        cliente = practica.cliente if not practica.cliente.baja else None
        if cliente and "GestionDeClientes.cliente_ver_habilitados" in poseidos:
            menu[1].append( (reverse("clientes:clienteVer", args=(cliente.id,)), "Ver cliente") )

    permitidas = filtrosPara(usuario, "listar")
    if permitidas is not None:
        menu[2].append(
            (reverse("facturas:listar"), "Listar facturas")
        )
    if "GestionDePagos.pago_listar" in poseidos:
        menu[2].append( (reverse("pagos:listar"), "Listar pagos") )
    return [ item for item in menu if len(item)]

def menuListar(usuario, accion):
    menu = [[], []]
    if accion != "exportar_xlsx":
        permitidas = filtrosPara(usuario, "exportar_xlsx")
        if permitidas is not None:
            menu[0].append(
                (reverse("facturas:exportar:xlsx"), "Exportar en xlsx")
            )
    if accion != "listar":
        permitidas = filtrosPara(usuario, "listar")
        if permitidas is not None:
            menu[1].append(
                (reverse("facturas:listar"), "Listar facturas")
            )
    if usuario.has_perm("GestionDePagos.pago_listar"):
        menu[1].append( (reverse("pagos:listar"), "Listar pagos") )
    return [ item for item in menu if len(item)]


def filtrosPara(usuario, accion):
    filtros = []
    retorno = None
    if usuario.has_perm("GestionDeFacturas.factura_%s_pagas" % accion):
        filtros.append( ~Q(pago=None) )
    if usuario.has_perm("GestionDeFacturas.factura_%s_no_pagas" % accion):
        filtros.append( Q(pago=None) )
    if len(filtros):
        if len(filtros) == 2:
            retorno = Q()
        else:
            retorno = filtros[0]
    return retorno



def listarXlsx(nombre, resultados, visibles, encabezados=None):

    libro = Workbook()
    hoja = libro.create_sheet(nombre, 0)
    fila = 1
    if encabezados:
        columna = 1
        for encabezado in encabezados:
            celda = hoja.cell(fila, columna, encabezado)
            columna += 1
        fila += 1
    for resultado in resultados:
        columna = 1
        if visibles[0]:
            celda = hoja.cell(fila, columna, resultado.id)
            columna += 1
        if visibles[1]:
            celda = hoja.cell(fila, columna, resultado.tipo)
            columna += 1
        if visibles[2]:
            celda = hoja.cell(fila, columna, resultado.total)
            columna += 1
        if visibles[3]:
            importe = None
            if resultado.estaPaga():
                importe = resultado.practica.precio
            celda = hoja.cell(fila, columna, importe)
            columna += 1
        if visibles[4]:
            importe = resultado.importe_ventas or 0
            celda = hoja.cell(fila, columna, importe)
            columna += 1
        if visibles[5]:
            cliente = resultado.cliente
            tostr = "%s: %s %s" % (cliente.dniCuit, cliente.apellidos, cliente.nombres)
            celda = hoja.cell(fila, columna, tostr)
            columna += 1
        if visibles[6]:
            celda = hoja.cell(fila, columna, resultado.fecha)
            columna += 1
        if visibles[7]:
            fecha = None
            if resultado.estaPaga():
                fecha = resultado.pago.fecha
            celda = hoja.cell(fila, columna, fecha)
            columna += 1
        fila += 1
    return guardarWorkbook(libro)


def crearFactura(form, usuario):
    if not form.is_valid():
        raise VeterinariaPatagonicaError("No se pudieron guardar los datos de la facturación, los datos del formulario no eran validos.")
    try:
        with transaction.atomic():
            factura = form.factura()
            factura.save(force_insert=True)
            productosFactura = form.productosFactura()
            if len(productosFactura)>0:
                factura.detalles_producto.set(productosFactura, bulk=False)
    except dbError as error:
        raise VeterinariaPatagonicaError({
            "titulo":"Error al guardar datos",
            "descripcion":"No se pudo guardar la factura en la base de datos."
        })

    practica = factura.practica
    if practica is not None:
        try:
            practica.hacer(usuario, Practica.Acciones.facturar.name)
        except:
            factura.delete()
            raise

    if factura.totalAdeudado() == 0:
        pago = Pago(factura=factura)
        try:
            pago.save(force_insert=True)
        except dbError as error:
            raise VeterinariaPatagonicaError({
                "titulo":"Error al guardar datos",
                "descripcion":"No se pudo guardar la factura en la base de datos."
            })
    return factura


def buscarPractica(id):
    practica = None
    try:
        practica = Practica.objects.get(id=id)
    except Practica.DoesNotExist:
        raise VeterinariaPatagonicaError("Error", "Práctica no encontrada")
    return practica


def facturarPractica(request, id):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    if not request.user.has_perm("GestionDeFacturas.factura_crear"):
        raise PermissionDenied()

    context = { "errores" : [], "accion" : None }

    try:
        practica = buscarPractica(id)
        verificarFacturacion(request.user, practica)
    except VeterinariaPatagonicaError as error:
        context["errores"].append(error)
    else:
        context["accion"] = "Guardar"
        context["menu"] = menuCrear(request.user, practica)
        context["facturacion"] = {
            "practica" : practicaRealizadaSelectLabel(practica),
            "cliente" : clienteSelectLabel(practica.cliente)
        }

        if request.method == "POST":
            acciones = AccionesFacturacionForm(request.POST)
            formset = ProductosFormSet(request.POST, prefix="producto")
            form = FacturarPracticaForm(practica, request.POST, formsetProductos=formset)

            if form.is_valid():
                if acciones.accion()=="guardar":
                    try:
                        factura = crearFactura(form, request.user)
                    except VeterinariaPatagonicaError as error:
                        context["errores"].append(error)
                    else:
                        return HttpResponseRedirect(
                            reverse("facturas:ver", args=(factura.id,))
                        )
        else:
            acciones = AccionesFacturacionForm()
            formset = ProductosFormSet(prefix="producto")
            form = FacturarPracticaForm(practica, formsetProductos=formset)

        context["form"]     = form
        context['formset']  = formset
        context['acciones']  = acciones

    template = loader.get_template("OtraGestionDeFacturas/facturarPractica.html")
    return HttpResponse(template.render(context, request))


def facturar(request):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    if not request.user.has_perm("GestionDeFacturas.factura_crear"):
        raise PermissionDenied()

    factura = None
    context = {
        "errores" : [],
        "accion" : "Guardar",
        "menu" : menuCrear(request.user)
    }

    if request.method == "POST":
        acciones = AccionesFacturacionForm(request.POST)
        formset = ProductosFormSet(request.POST, prefix="producto")
        form = FacturarForm(request.POST, formsetProductos=formset)

        if form.is_valid():
            factura = form.factura()
            productos = form.productosFactura()
            practica = factura.practica

            try:
                if practica is not None:
                    verificarFacturacion(request.user, practica)
            except VeterinariaPatagonicaError as error:
                context["errores"].append(error)
            else:
                if acciones.accion()=="guardar":
                    try:
                        factura = crearFactura(form, request.user)
                    except VeterinariaPatagonicaError as error:
                        context["errores"].append(error)
                    else:
                        return HttpResponseRedirect(
                            reverse("facturas:ver", args=(factura.id,))
                        )
    else:
        acciones = AccionesFacturacionForm()
        formset = ProductosFormSet(prefix="producto")
        form = FacturarForm(formsetProductos=formset)

    context["form"]     = form
    context['formset']  = formset
    context['acciones']  = acciones

    template = loader.get_template("OtraGestionDeFacturas/facturar.html")
    return HttpResponse(template.render(context, request))


def listar(request):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    permitidas = filtrosPara(request.user, "listar")
    if permitidas is None:
        raise PermissionDenied()
    facturas = Factura.objects.filter(permitidas)

    gestor = GestorListadoFacturas(
        campos = [
            ["id", "Número"],
            ["tipo", "Tipo"],
            ["total", "Importe"],
            ["precioPractica", "Precio de practica"],
            ["importeVentas", "Importe por ventas"],
            ["cliente", "Cliente"],
            ["fecha", "Fecha"],
            ["fechaPago", "Fecha de pago"],
        ],
        iniciales={
            "seleccion" : (
                "tipo",
                "total",
                "cliente",
                "fecha",
            ),
            "paginacion" : {
                "cantidad" : 10,
            },
            "orden" : (
                ("fecha", False),
                ("total", False),
                ("tipo", True),
            ),
        },
        clases={ "filtrado" : FiltradoFacturaForm },
        queryset=facturas,
        mapaFiltrado=facturas.MAPEO_FILTRADO,
        mapaOrden=facturas.MAPEO_ORDEN,
    )

    gestor.cargar(request)
    gestor.paginacion["cantidad"].label = "Facturas por pagina"

    verPagas = request.user.has_perm("GestionDeFacturas.factura_ver_pagas")
    verNoPagas = request.user.has_perm("GestionDeFacturas.factura_ver_no_pagas")
    visibles = set()
    for factura in gestor.itemsActuales():
        if factura.estaPaga():
            if verPagas:
                visibles.add(factura.id)
        else:
            if verNoPagas:
                visibles.add(factura.id)

    template = loader.get_template("OtraGestionDeFacturas/listar.html")
    context = {
        "gestor" : gestor,
        "visibles" : visibles,
        "menu" : menuListar(request.user, "listar"),
    }

    return HttpResponse(template.render( context, request ))

def ver(request, id):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    factura = Factura.objects.get(id=id)
    condicion = "pagas" if factura.estaPaga() else "no_pagas"
    if not request.user.has_perm("GestionDeFacturas.factura_ver_%s" % condicion):
        raise PermissionDenied()

    practica = factura.practica or None
    permisoVer = None
    if practica is not None:
        permisoVer =  permisos.paraPractica(
            request.user,
            Practica.Acciones.ver_informacion_general,
            practica
        )

    context = {
        "factura" : factura,
        "urlPractica" : pathVer(
            practica.nombreTipo(),
            practica.id
        ) if permisoVer else "",
        "menu" : menuVer(request.user, factura)
    }

    template = loader.get_template("OtraGestionDeFacturas/ver.html")
    return HttpResponse(template.render( context, request ))

def exportar(request, formato=None):

    if isinstance(request.user, AnonymousUser):
        return HttpResponseRedirect("%s?proxima=%s" % (LOGIN_URL, request.path))

    permitidas = filtrosPara(request.user, "exportar_%s" % formato)
    if permitidas is None:
        raise PermissionDenied()
    facturas = Factura.objects.filter(permitidas)

    gestor = GestorListadoFacturas(
        campos = [
            ["id", "Número"],
            ["tipo", "Tipo"],
            ["total", "Importe"],
            ["precioPractica", "Precio de practica"],
            ["importeVentas", "Importe por ventas"],
            ["cliente", "Cliente"],
            ["fecha", "Fecha"],
            ["fechaPago", "Fecha de pago"],
        ],
        iniciales={
            "seleccion" : (
                "tipo",
                "total",
                "cliente",
                "fecha",
            ),
            "paginacion" : {
                "cantidad" : 25,
            },
        },
        clases={ "filtrado" : FiltradoFacturaForm },
        queryset=facturas,
        mapaFiltrado=facturas.MAPEO_FILTRADO,
        mapaOrden=facturas.MAPEO_ORDEN,
    )

    gestor.cargar(request)
    gestor.paginacion["cantidad"].label = "Facturas por pagina"
    exportar = ExportarForm(request.GET)
    accion = exportar.accion()

    if accion == "":
        template = loader.get_template("OtraGestionDeFacturas/exportar.html")
        context = {
            "formato" : formato,
            "gestor" : gestor,
            "exportar" : exportar,
            "menu" : menuListar(request.user, "exportar_"+formato),
        }
        retorno = HttpResponse(template.render(context, request))
    else:
        nombre = "facturas-%s" % datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        visibles = [ gestor[campo]["visible"] for campo in gestor.columnas ]
        encabezados = [ gestor[campo]["etiqueta"] for campo in gestor.columnasVisibles ]

        if accion=="exportar_pagina":
            facturas = gestor.itemsActuales()
        elif accion=="exportar_todos":
            facturas = gestor.items()

        contenido = listarXlsx("facturas", facturas, visibles, encabezados)
        retorno = HttpResponse(contenido, content_type="application/ms-excel")
        retorno["Content-Disposition"] = "attachment; filename=%s.xlsx" % nombre
    return retorno


def ayudaContextualCosto(request):
# Redireccionamos la ayuda contextual
    template = loader.get_template('GestionDeFacturas/ayudaGestiondeCostos.html')
    contexto = {
        'usuario': request.user,
    }
    return HttpResponse(template.render(contexto, request))
    